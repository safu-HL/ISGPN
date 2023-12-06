from openpyxl import load_workbook

#---------------------------------读取NMF 关键生物标志物
# K=39
for K in range(6,13):
    mRNA_index=[]
    miRNA_index=[]
    lncRNA_index=[]
    RNA_index=[mRNA_index,miRNA_index,lncRNA_index]
    RNA_name=["mRNA","miRNA","lncRNA"]
    for rna in range(3):
        tt=load_workbook("D:/matlabfiles/K-NMF RNA/tt/Co_module_tt_2_K_42_svd_"+RNA_name[rna]+"_2.xlsx")
        ttTable=tt.get_sheet_by_name(tt.get_sheet_names()[0])
        i=1
        while str(ttTable.cell(i,K).value)!="None":
            RNA_index[rna].append(int(ttTable.cell(i,K).value))
            i+=1
        tt.close()

    #------------------------------------读取名字
    mRNA_index_name=[]
    miRNA_index_name=[]
    lncRNA_index_name=[]
    RNA_index_name=[mRNA_index_name,miRNA_index_name,lncRNA_index_name]
    for rna in range(3):
        gene_data=load_workbook("D:/pythonfiles/bio information/NMF RNA/rna/"+RNA_name[rna]+"_names.xlsx")
        gene_dataTable=gene_data.get_sheet_by_name(gene_data.get_sheet_names()[0])
        maxrow=gene_dataTable.max_row
        maxcol=gene_dataTable.max_column
        for i in range(maxrow - 1):
            if i+1 in RNA_index[rna]:
                RNA_index_name[rna].append(str(gene_dataTable.cell(i+2,1).value))
        gene_data.close()

    #-----------------------------------生成数据
    interaction=set()

    for i in lncRNA_index_name:
        for j in miRNA_index_name:
            interaction.add(i+","+j)
    for i in miRNA_index_name:
        for j in mRNA_index_name:
            interaction.add(i+","+j)
    print(interaction)
    print(len(interaction))
    #------------------------------过滤ceRNA
    input2=open("../data/input2.txt","r")
    line=input2.readline()
    while line:
        a, b = line.split("\t")
        interaction.add(a+","+b[0:-1])
        line=input2.readline()
    input2.close()
    print(interaction)
    print(len(interaction))

    #-----------------write
    input_interaction=open("../data/"+str(K)+"/input_interaction_"+str(K)+".txt","w")
    for i in interaction:
        a,b=i.split(",")
        input_interaction.write(a+"\t"+b+"\n")
    input_interaction.close()





