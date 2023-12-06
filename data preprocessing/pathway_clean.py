from openpyxl import load_workbook
import pandas as pd
import numpy as np

# K=39
for K in range(6,13):
    mrna=load_workbook("../data/"+str(K)+"/mRNA_"+str(K)+".xlsx")
    mrnaTable=mrna.get_sheet_by_name(mrna.get_sheet_names()[0])
    maxcol=mrnaTable.max_column
    mrna_clean=[]
    for i in range(maxcol-1):
        mrna_clean.append(mrnaTable.cell(1,i+2).value)
    mrna.close()

    print(mrna_clean)
    print(len(mrna_clean))

    path = pd.read_csv("D:/Shanghai Maritime/LUSC2/pathway/pathwayMask1.csv",index_col=0)
    name=path.columns.values.tolist()

    select_rna=set(mrna_clean)&set(name)
    print(select_rna)
    print(len(select_rna))

    print(path)
    path=path.loc[:,select_rna]
    print(path)

    delete_index=[]
    for i in range(path.shape[0]):
        temp=np.array(path.iloc[i].values.tolist())
        if temp.sum==0:
            delete_index.append(i)
    print(delete_index)
    print(len(delete_index))

    path=path.drop(delete_index)
    print(path)

    path.to_csv("../data/"+str(K)+"/pathwayMask_"+str(K)+".csv")