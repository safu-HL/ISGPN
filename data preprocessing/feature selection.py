from openpyxl import load_workbook
import xlsxwriter as xw
import pandas as pd
#-------------------------读取ceRNA信息
mRNA=set()
miRNA=set()
lncRNA=set()
RNA=[mRNA,miRNA,lncRNA]
input2=open("../data/input2.txt","r")
line=input2.readline()
while line:
    a,b=line.split("\t")
    if a.split("-")[0]=="hsa":
        miRNA.add(a)
        mRNA.add(b[0:-1])
    else:
        lncRNA.add(a)
        miRNA.add(b[0:-1])
    line=input2.readline()

#---------------------------------读取NMF 关键生物标志物
# K=39
for K in range(6,13):
    print(K)
    mRNA_index=[]
    miRNA_index=[]
    lncRNA_index=[]
    RNA_index=[mRNA_index,miRNA_index,lncRNA_index]
    RNA_name=["mRNA","miRNA","lncRNA"]

    for rna in range(3):
        tt=load_workbook("D:/matlabfiles/K-NMF RNA/tt/Co_module_tt_2_K_42_svd_"+RNA_name[rna]+"_2.xlsx")
        ttTable=tt.get_sheet_by_name(tt.get_sheet_names()[0])
        i=1
        while str(ttTable.cell(i,K).value)!="None":
            RNA_index[rna].append(int(ttTable.cell(i,K).value))
            i+=1
        tt.close()

    #--------------------------------生成数据

    for rna in range(3):
        gene_raw=xw.Workbook("../data/"+str(K)+"/"+RNA_name[rna]+"_"+str(K)+".xlsx")
        gene_rawTable=gene_raw.add_worksheet("sheet1")
        gene_rawTable.activate()

        gene_data=load_workbook("D:/pythonfiles/bio information/NMF RNA/rna/"+RNA_name[rna]+"_names.xlsx")
        gene_dataTable=gene_data.get_sheet_by_name(gene_data.get_sheet_names()[0])
        maxrow=gene_dataTable.max_row
        maxcol=gene_dataTable.max_column

        for i in range(maxcol-1):
            gene_rawTable.write(i+1,0,gene_dataTable.cell(1,i+2).value)

        select_geneindex=0
        for i in range(maxrow-1):
            if str(gene_dataTable.cell(i+2,1).value) in RNA[rna] or i+1 in RNA_index[rna]:
                select_geneindex+=1
                gene_rawTable.write(0,select_geneindex,gene_dataTable.cell(i+2,1).value)
                for j in range(maxcol-1):
                    gene_rawTable.write(j+1,select_geneindex,float(gene_dataTable.cell(i+2,j+2).value))
        gene_raw.close()
        gene_data.close()
        excelfile=pd.ExcelFile("../data/"+str(K)+"/"+RNA_name[rna]+"_"+str(K)+".xlsx")
        df=excelfile.parse(excelfile.sheet_names[0])
        df.to_csv("../data/"+str(K)+"/"+RNA_name[rna]+"_"+str(K)+".csv",index=False)






