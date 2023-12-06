import copy
import numpy as np
import torch
import pandas as pd
import random
import matplotlib.pyplot as plt
import torch.nn.functional as F
import torch.nn as nn
from openpyxl import load_workbook
import xlsxwriter as xw
import dgl

# RNA_type={}
# mRNA=load_workbook("F:/Shanghai Maritime University/LUSC/RNA-seq/05/DEmRNA.xlsx")
# mRNA_table=mRNA.get_sheet_by_name(mRNA.get_sheet_names()[0])
# row=mRNA_table.max_row
# for i in range(row):
#     temp=mRNA_table.cell(i + 1, 1).value.split("|")[0]
#     if float(mRNA_table.cell(i+1,3).value)>0:
#         RNA_type[temp]="mRNAdown"
#     else:
#         RNA_type[temp] = "mRNAup"
# mRNA.close()
#
# lncRNA=load_workbook("F:/Shanghai Maritime University/LUSC/RNA-seq/05/DElncRNA.xlsx")
# lncRNA_table=lncRNA.get_sheet_by_name(lncRNA.get_sheet_names()[0])
# row=lncRNA_table.max_row
# for i in range(row):
#     temp=lncRNA_table.cell(i+1,1).value.split("|")[0]
#     if float(lncRNA_table.cell(i+1,3).value)>0:
#         RNA_type[temp]="lncRNAdown"
#     else:
#         RNA_type[temp]="lncRNAup"
# lncRNA.close()
#
# miRNA=load_workbook("F:/Shanghai Maritime University/LUSC/RNA-seq/05/DEmiRNA.xlsx")
# miRNA_table=miRNA.get_sheet_by_name(miRNA.get_sheet_names()[0])
# row=miRNA_table.max_row
# for i in range(row):
#     temp=miRNA_table.cell(i+1,1).value
#     if float(miRNA_table.cell(i+1,3).value)>0:
#         RNA_type[temp]="miRNAdown"
#     else:
#         RNA_type[temp]="miRNAup"
# miRNA.close()
#
# file=open("F:/python file/bio network/ceGNN2/data/5/type.txt","w")
#
# gene_set=set()
# connection=load_workbook("F:/python file/bio network/ceGNN2/data/5/new_ceRNA_5_backup.xlsx")
# connection_table=connection.get_sheet_by_name(connection.get_sheet_names()[0])
# row=connection_table.max_row
# index=1
# for i in range(row):
#     if 15<=index<=83:
#         temp=connection_table.cell(index,1).value
#         if(temp[-1]=="\n"):
#             gene_set.add(connection_table.cell(index,1).value[0:-1])
#         else:
#             gene_set.add(connection_table.cell(index, 1).value)
#         temp = connection_table.cell(index, 2).value
#         if(temp[-1]=="\n"):
#             gene_set.add(connection_table.cell(index,2).value[0:-1])
#         else:
#             gene_set.add(connection_table.cell(index, 2).value)
#     index+=1
# connection.close()
#
# for i in gene_set:
#     print(i)
#     file.write(i+"\t"+RNA_type[i]+"\n")
# file.close()

corrlist=pd.read_csv("data/5/m_path_sc_backup.csv",index_col=0)

gene_name = corrlist.columns.values.tolist()
path_name = corrlist.index.values.tolist()
print(gene_name)

print(corrlist.values)
